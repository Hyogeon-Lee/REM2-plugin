"""XLSX → Markdown 변환 어댑터.

전략:
- 시트별 ## 헤더
- 사용된 영역(used range)만 마크다운 표로 변환
- 이미지가 있으면 해당 시트의 표 아래에 anchor 셀 정보와 함께 삽입
  (시트 안에서 위치 매칭은 행 단위로 근사 — XLSX는 본문이 아니라 그리드라 정확한 inline 위치는 의미가 약함)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..common import ImageSaver, make_image_markdown


def convert(input_path: Path, output_md_path: Path) -> dict:
    wb = load_workbook(str(input_path), data_only=True)
    saver = ImageSaver(output_md_path, output_md_path.stem, image_dir_name="images")
    md_lines: list[str] = [f"# {output_md_path.stem}"]
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"\n## {sheet_name}")

        table_md = _sheet_to_table(ws)
        if table_md:
            md_lines.append(table_md)

        # 이미지: openpyxl은 ws._images에 보관 (private이지만 표준 접근 방식)
        images = list(getattr(ws, "_images", []))
        if images:
            md_lines.append("\n**이미지:**\n")
            for img in images:
                blob = _extract_image_blob(img)
                if blob is None:
                    warnings.append(f"시트 {sheet_name}: 이미지 추출 실패")
                    continue
                anchor = _describe_anchor(img)
                rel_path = saver.save(blob, ext=_guess_ext(img))
                tag = make_image_markdown(rel_path, alt=anchor or "image")
                md_lines.append(f"{tag}  <!-- {anchor} -->" if anchor else tag)

    output_md_path.write_text("\n\n".join(md_lines) + "\n", encoding="utf-8")
    warnings.extend(saver.warnings)
    return {"image_count": saver.saved_count, "warnings": warnings}


def _sheet_to_table(ws) -> str:
    """시트의 사용된 영역을 마크다운 표로."""
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
        return ""

    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                cells.append(" ")
            else:
                cells.append(str(v).replace("\n", "<br>").replace("|", r"\|"))
        rows.append(cells)

    # 완전히 빈 행 끝부분 제거
    while rows and all(c.strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        return ""

    n_cols = max(len(r) for r in rows)
    rows = [r + [" "] * (n_cols - len(r)) for r in rows]

    out = ["| " + " | ".join(rows[0]) + " |", "|" + "|".join(["---"] * n_cols) + "|"]
    for r in rows[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def _extract_image_blob(img) -> bytes | None:
    """openpyxl Image 객체에서 blob 추출."""
    # openpyxl 버전에 따라 _data 또는 ref 접근 방식 다름
    try:
        if hasattr(img, "_data") and callable(img._data):
            return img._data()
        if hasattr(img, "ref"):
            ref = img.ref
            if hasattr(ref, "read"):
                ref.seek(0)
                return ref.read()
            if isinstance(ref, (str, Path)):
                return Path(ref).read_bytes()
    except Exception:
        return None
    return None


def _guess_ext(img) -> str:
    """이미지 객체의 format 속성에서 확장자 추론."""
    fmt = getattr(img, "format", None)
    if fmt:
        return str(fmt).lower()
    return "png"


def _describe_anchor(img) -> str:
    """이미지의 셀 anchor 위치를 사람이 읽을 수 있는 문자열로."""
    try:
        anchor = img.anchor
        if hasattr(anchor, "_from"):
            f = anchor._from
            return f"anchor: row={f.row + 1}, col={f.col + 1}"
    except Exception:
        pass
    return ""
